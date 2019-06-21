import os
import csv
import pandas as pd
import openpyxl
from textblob import TextBlob
from openpyxl import load_workbook
from itertools import islice
import nltk
from pandas import read_excel
os.environ['STANFORD_PARSER'] = 'D:\Subjects\dissertation\TripleExtractionProject\stanford-parser-full-2018-10-17\stanford-parser-full-2018-10-17\stanford-parser'
os.environ['STANFORD_MODELS'] = 'D:\Subjects\dissertation\TripleExtractionProject\stanford-parser-full-2018-10-17\stanford-parser-full-2018-10-17\stanford-parser-3.9.2-models'
from nltk.parse.stanford import StanfordParser
from nltk.tree import ParentedTree, Tree
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

file = "D:\Subjects\dissertation\DATASET\\test.xlsx"
wb = load_workbook(file)
sheet = wb.get_sheet_by_name('Sheet1')
dataframe1 = sheet.values


cols = next(dataframe1)[1:]
dataframe1 = list(dataframe1)
idx = [r[0] for r in dataframe1]
dataframe1 = (islice(r,1,None)for r in dataframe1)

df = pd.DataFrame(dataframe1,index=idx,columns=cols)

columns = list(df.columns.values)
parser = StanfordParser()

#getting all the values from excel sheet to variable 'sent'
for col in df.columns[2:3]:
    sent = df[col]
    
#for storing the result
wb=Workbook()
wb.save('D:\Subjects\dissertation\DATASET\\russu.xlsx')
sheet1 = wb['Sheet']
df2 = pd.DataFrame()


def find_subject(t):
    for s in t.subtrees(lambda t: t.label() == 'NP'):
        for n in s.subtrees(lambda n: n.label().startswith('NN')):
            return (n[0], find_attrs(n))


def find_predicate(t):
    v = None

    for s in t.subtrees(lambda t: t.label() == 'VP'):
        for n in s.subtrees(lambda n: n.label().startswith('VB')):
            v = n
        return (v[0], find_attrs(v))


def find_object(t):
    for s in t.subtrees(lambda t: t.label() == 'VP'):
        for n in s.subtrees(lambda n: n.label() in ['NP', 'PP', 'ADJP']):
            if n.label() in ['NP', 'PP']:
                for c in n.subtrees(lambda c: c.label().startswith('NN')):
                    return (c[0], find_attrs(c))
            else:
                for c in n.subtrees(lambda c: c.label().startswith('JJ')):
                    return (c[0], find_attrs(c))


def find_attrs(node):
    attrs = []
    p = node.parent()

    # Search siblings
    if node.label().startswith('JJ'):
        for s in p:
            if s.label() == 'RB':
                attrs.append(s[0])

    elif node.label().startswith('NN'):
        for s in p:
            if s.label() in ['DT', 'PRP$', 'POS', 'JJ', 'CD', 'ADJP', 'QP', 'NP']:
                attrs.append(' '.join(s.flatten()))

    elif node.label().startswith('VB'):
        for s in p:
            if s.label() == 'ADVP':
                attrs.append(' '.join(s.flatten()))

        # Search uncles
    if node.label().startswith('JJ') or node.label().startswith('NN'):
            for s in p.parent():
                if s != p and s.label() == 'PP':
                    attrs.append(' '.join(s.flatten()))

    elif node.label().startswith('VB'):
            for s in p.parent():
                if s != p and s.label().startswith('VB'):
                    attrs.append(s[0])

    return attrs

for j in sent.values:
    if(j!=None):
        t = list(parser.raw_parse(j))[0]
        t = ParentedTree.convert(t)
        #t.pretty_print()
        print(find_subject(t))
        print(find_predicate(t))
        print(find_object(t))
        # create a panda dataframe to store the subject predicate and object
        df = pd.DataFrame({'Subject': [find_subject(t)],
                           'Predicate': [find_predicate(t)],
                           'Object': [find_predicate(t)],
                           'Question':[j]})

        df2=df2.append(df,ignore_index=True)
        df2[df2.Subject != '[]']
        df2.to_excel(r'D:\Subjects\dissertation\DATASET\\russu.xlsx')



